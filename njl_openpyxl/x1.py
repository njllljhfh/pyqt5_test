# -*- coding:utf-8 -*-
from openpyxl import Workbook

wb = Workbook()

# 获得激活的worksheet
ws = wb.active
ws.title = 'x1'

# 数据可以直接赋值给单元格
ws['A1'] = 42

# 也可以按行添加
ws.append([1, 2, 3])  # 在当前文档的最后一行下面添加一行数据
ws.append([4, 5, 6])

# Python类型将自动转换
import datetime

ws['A2'].number_format = 'yyyy年mm月dd日'
ws['A2'] = datetime.datetime.now()

ws3 = wb.create_sheet("x3")  # 在末尾插入(default)
ws2 = wb.create_sheet("x2", 0)  # 在最前端插入
ws4 = wb.create_sheet("x4", -1)  # 插入在倒数第二个位置
# ws4.sheet_properties.tabColor = "1072BA"

# 一旦你给工作表(worksheet)命名, 你就可以通过workbook的key来获取它:
x4 = wb['x4']
x4.sheet_properties.tabColor = "FF0000"

# 你可以通过 Workbook.sheetname属性来重新审阅工作表的名称
print(wb.sheetnames)
print('- ' * 30)

# 你可以循环浏览工作表
for sheet in wb:
    print(sheet.title)

# 您可以在 单个工作簿中创建工作表的副本:
source = wb['x1']
target = wb.copy_worksheet(source)

# 保存文件
wb.save("sample.xlsx")
